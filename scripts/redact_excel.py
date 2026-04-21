#!/usr/bin/env python3
"""
Redact Excel (.xlsx): strips embedded images/charts and contact-identifying cell
content (supplier OR buyer, depending on CONTACT_KIND) via the OpenAI API (first
20 rows per sheet sampled for AI context).

Usage: reads raw xlsx bytes from stdin, writes cleaned xlsx bytes to stdout.
  - OPENAI_API_KEY, OPENAI_BASE_URL, BYOS_AI_MODEL: AI config (from env)
  - CONTACT_ROSTER: JSON array of {canonicalName, aliases} (from env, set by TS caller)
  - CONTACT_KIND: "supplier" or "buyer" (controls the system prompt only).
  - REDACTION_LABEL: string to substitute for identifying mentions; defaults
    to the generic marker "[REDACTED]" used by the rest of the BYOS redaction
    pipeline (see byos/src/redact.ts — we deliberately do NOT distinguish
    [REDACTED SUPPLIER] vs [REDACTED BUYER] since nothing downstream keys off
    the distinction and a shorter token saves prompt tokens on every
    subsequent LLM pass over the redacted body).

Exit code 0 on success, non-zero on failure (stderr has the error message).
"""
import json
import os
import sys
from io import BytesIO

from openpyxl import load_workbook

SAMPLE_ROWS = 20


def get_openai_client():
    api_key = os.environ.get("OPENAI_API_KEY", "")
    if not api_key:
        return None
    from openai import OpenAI

    base_url = os.environ.get("OPENAI_BASE_URL") or None
    return OpenAI(api_key=api_key, base_url=base_url)


def get_contact_kind() -> str:
    kind = (os.environ.get("CONTACT_KIND") or "supplier").strip().lower()
    return "buyer" if kind == "buyer" else "supplier"


def get_redaction_label() -> str:
    explicit = os.environ.get("REDACTION_LABEL")
    if explicit:
        return explicit
    return "[REDACTED]"


def get_contact_roster() -> list[dict]:
    raw = os.environ.get("CONTACT_ROSTER", "")
    if not raw:
        return []
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        print("Warning: failed to parse CONTACT_ROSTER", file=sys.stderr)
        return []


def sample_sheet_content(ws, max_rows: int = SAMPLE_ROWS) -> str:
    """Extract the first N rows as a TSV-like string for AI context."""
    lines = []
    for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True), start=1):
        cells = [str(c) if c is not None else "" for c in row]
        if any(cells):
            lines.append(f"Row {row_idx}: " + "\t".join(cells))
    return "\n".join(lines)


def ask_ai_for_redactions(
    client,
    model: str,
    roster: list[dict],
    sheet_samples: dict[str, str],
    kind: str,
    label: str,
) -> list[dict]:
    """
    Send sampled sheet content + contact roster to the AI.
    Returns a list of {needle, replacement} redaction rules.
    """
    roster_summary = json.dumps(
        [{"name": s.get("canonicalName", ""), "aliases": s.get("aliases", [])} for s in roster],
        ensure_ascii=False,
    )

    sheets_text = "\n\n".join(
        f"=== Sheet: {name} ===\n{content}" for name, content in sheet_samples.items() if content
    )

    if not sheets_text.strip():
        return []

    if kind == "buyer":
        system_prompt = (
            f"You redact BUYER-identifying information from spreadsheet content. "
            f"The buyer is the customer/client the request is for, NOT the sender (the sender is "
            f"typically an internal employee relaying a request). You receive a buyer roster and "
            f"sampled rows from Excel sheets. Return a JSON object with a single key \"redactions\": "
            f"an array of {{\"needle\": \"<exact case-sensitive substring from the input>\", "
            f"\"replacement\": \"{label}\"}} objects. Each needle must be a literal substring copied "
            f"verbatim from the sheet content. Redact: buyer/customer/company names, contact persons, "
            f"phone numbers, email addresses, physical addresses, website URLs, and any other "
            f"identifying information that reveals who the buyer is. Do NOT redact product names, "
            f"SKUs, quantities, prices, units, or other trade-relevant data. Return an empty "
            f"redactions array if nothing needs redacting."
        )
    else:
        system_prompt = (
            f"You redact supplier-identifying information from spreadsheet content. "
            f"You receive a supplier roster and sampled rows from Excel sheets. "
            f"Return a JSON object with a single key \"redactions\": an array of "
            f"{{\"needle\": \"<exact case-sensitive substring from the input>\", "
            f"\"replacement\": \"{label}\"}} objects. Each needle must be a literal substring copied "
            f"verbatim from the sheet content. Redact: supplier/company names, contact persons, phone "
            f"numbers, email addresses, physical addresses, website URLs, and any other identifying "
            f"information that reveals who the supplier is. Do NOT redact product names, SKUs, "
            f"quantities, prices, units, or other trade-relevant data. Return an empty redactions "
            f"array if nothing needs redacting."
        )

    response = client.chat.completions.create(
        model=model,
        response_format={"type": "json_object"},
        messages=[
            {"role": "system", "content": system_prompt},
            {
                "role": "user",
                "content": json.dumps(
                    {"kind": kind, "roster": roster_summary, "sheets": sheets_text},
                    ensure_ascii=False,
                ),
            },
        ],
    )

    raw = response.choices[0].message.content
    if not raw:
        return []
    try:
        parsed = json.loads(raw)
        redactions = parsed.get("redactions", [])
        return [
            r for r in redactions
            if isinstance(r, dict)
            and isinstance(r.get("needle"), str)
            and isinstance(r.get("replacement"), str)
            and r["needle"]
            and r["needle"] != r["replacement"]
        ]
    except (json.JSONDecodeError, AttributeError):
        print("Warning: failed to parse AI redaction response", file=sys.stderr)
        return []


def apply_redactions(wb, redactions: list[dict]) -> int:
    """Apply needle→replacement substitutions across all cells in the workbook. Returns count of changes."""
    if not redactions:
        return 0
    count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                original = cell.value
                for r in redactions:
                    if r["needle"] in cell.value:
                        cell.value = cell.value.replace(r["needle"], r["replacement"])
                if cell.value != original:
                    count += 1
    return count


def strip_and_redact(data: bytes) -> bytes:
    wb = load_workbook(BytesIO(data))

    # Strip images and charts
    for ws in wb.worksheets:
        ws._images = []
        ws._charts = []

    # AI-based content redaction
    client = get_openai_client()
    if client:
        model = os.environ.get("BYOS_AI_MODEL", "gpt-4.1-mini")
        kind = get_contact_kind()
        label = get_redaction_label()
        roster = get_contact_roster()
        sheet_samples = {}
        for ws in wb.worksheets:
            content = sample_sheet_content(ws)
            if content:
                sheet_samples[ws.title] = content
        if sheet_samples:
            try:
                redactions = ask_ai_for_redactions(
                    client, model, roster, sheet_samples, kind, label
                )
                changed = apply_redactions(wb, redactions)
                if changed:
                    print(
                        f"Redacted {changed} {kind} cell(s) across "
                        f"{len(wb.worksheets)} sheet(s)",
                        file=sys.stderr,
                    )
            except Exception as exc:
                print(f"Warning: AI redaction failed, images still stripped: {exc}", file=sys.stderr)
    else:
        print("No OPENAI_API_KEY set, skipping content redaction (images still stripped)", file=sys.stderr)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def main() -> None:
    raw = sys.stdin.buffer.read()
    if not raw:
        sys.exit("No input received on stdin")
    try:
        cleaned = strip_and_redact(raw)
    except Exception as exc:
        sys.exit(f"Failed to process Excel file: {exc}")
    sys.stdout.buffer.write(cleaned)


if __name__ == "__main__":
    main()
